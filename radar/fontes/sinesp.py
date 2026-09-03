"""Ocorrências criminais do SINESP, do Ministério da Justiça.

A planilha é nacional e tem 206 MB descomprimidos, então é lida em fluxo, uma
linha por vez, em vez de carregada na memória. Dos 31 eventos publicados, só 11
vêm com município; os outros existem apenas no total do estado.
"""

from typing import NamedTuple

import openpyxl

from radar.municipios import MunicipioDesconhecido, Sentinela, por_nome

MUNICIPALIZADOS = {
    "Feminicídio",
    "Homicídio doloso",
    "Lesão corporal seguida de morte",
    "Mandado de prisão cumprido",
    "Morte no trânsito ou em decorrência dele (exceto homicídio doloso)",
    "Mortes a esclarecer (sem indício de crime)",
    "Mortes no trânsito",
    "Roubo seguido de morte (latrocínio)",
    "Suicídio",
    "Tentativa de feminicídio",
    "Tentativa de homicídio",
}
BASE = (
    "https://www.gov.br/mj/pt-br/assuntos/sua-seguranca/seguranca-publica"
    "/estatistica/download/dnsp-base-de-dados"
)
UF, MUNICIPIO, EVENTO, DATA = 0, 1, 2, 3
TOTAL_VITIMA, TOTAL, ABRANGENCIA = 10, 11, 13


class Ocorrencia(NamedTuple):
    codigo_ibge: str
    ano: int
    mes: int
    evento: str
    abrangencia: str
    vitimas: int


def le_planilha(caminho, uf: str = "GO") -> list[Ocorrencia]:
    livro = openpyxl.load_workbook(caminho, read_only=True)
    try:
        pagina = livro[livro.sheetnames[0]]
        linhas = pagina.iter_rows(values_only=True)
        next(linhas)
        achadas = []
        for linha in linhas:
            if linha[UF] != uf or linha[EVENTO] not in MUNICIPALIZADOS:
                continue
            try:
                codigo = por_nome(linha[MUNICIPIO] or "")
            except (MunicipioDesconhecido, Sentinela):
                continue
            # a fonte preenche total_vitima ou total, nunca os dois
            vitimas = linha[TOTAL_VITIMA]
            if vitimas is None:
                vitimas = linha[TOTAL]
            data = linha[DATA]
            achadas.append(
                Ocorrencia(
                    codigo,
                    data.year,
                    data.month,
                    linha[EVENTO],
                    linha[ABRANGENCIA],
                    int(vitimas or 0),
                )
            )
        return achadas
    finally:
        livro.close()


def url_planilha(ano: int) -> str:
    return f"{BASE}/bancovde-{ano}.xlsx/@@download/file"
